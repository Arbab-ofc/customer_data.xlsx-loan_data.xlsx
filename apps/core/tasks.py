from celery import shared_task
from django.conf import settings
import openpyxl
from datetime import datetime
from decimal import Decimal
from apps.customers.models import Customer
from apps.loans.models import Loan


@shared_task(bind=True, max_retries=3)
def ingest_customer_data(self):
    """
    Background task to ingest customer data from Excel file
    """
    file_path = settings.DATA_DIR / 'customer_data.xlsx'

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        customers_created = 0
        customers_updated = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            customer_id = row[0]
            first_name = row[1]
            last_name = row[2]
            phone_number = row[3]
            monthly_salary = Decimal(str(row[4]))
            approved_limit = Decimal(str(row[5]))
            current_debt = Decimal(str(row[6])) if row[6] else Decimal('0')

            customer, created = Customer.objects.update_or_create(
                customer_id=customer_id,
                defaults={
                    'first_name': first_name,
                    'last_name': last_name,
                    'phone_number': phone_number,
                    'monthly_salary': monthly_salary,
                    'approved_limit': approved_limit,
                    'current_debt': current_debt
                }
            )

            if created:
                customers_created += 1
            else:
                customers_updated += 1

        return {
            'status': 'success',
            'customers_created': customers_created,
            'customers_updated': customers_updated
        }

    except Exception as e:
        self.retry(exc=e, countdown=60)


@shared_task(bind=True, max_retries=3)
def ingest_loan_data(self):
    """
    Background task to ingest loan data from Excel file
    """
    file_path = settings.DATA_DIR / 'loan_data.xlsx'

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        loans_created = 0
        loans_updated = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            customer_id = row[0]
            loan_id = row[1]
            loan_amount = Decimal(str(row[2]))
            tenure = int(row[3])
            interest_rate = Decimal(str(row[4]))
            monthly_repayment = Decimal(str(row[5]))
            emis_paid_on_time = int(row[6])

            start_date = row[7]
            if isinstance(start_date, str):
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()

            end_date = row[8]
            if isinstance(end_date, str):
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

            try:
                customer = Customer.objects.get(customer_id=customer_id)
            except Customer.DoesNotExist:
                continue

            loan, created = Loan.objects.update_or_create(
                loan_id=loan_id,
                defaults={
                    'customer': customer,
                    'loan_amount': loan_amount,
                    'tenure': tenure,
                    'interest_rate': interest_rate,
                    'monthly_repayment': monthly_repayment,
                    'emis_paid_on_time': emis_paid_on_time,
                    'start_date': start_date,
                    'end_date': end_date,
                    'is_active': True
                }
            )

            if created:
                loans_created += 1
            else:
                loans_updated += 1

        return {
            'status': 'success',
            'loans_created': loans_created,
            'loans_updated': loans_updated
        }

    except Exception as e:
        self.retry(exc=e, countdown=60)


@shared_task
def ingest_all_data():
    """
    Master task to ingest both customer and loan data
    """
    ingest_customer_data.delay()
    ingest_loan_data.delay()
    return {'status': 'Data ingestion tasks queued'}
